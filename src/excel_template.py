#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel-Template-Generator für PDFStructure2Excel
Erstellt eine Excel-Vorlage mit formatiertem Layout für die Strukturdaten
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_template(data, output_path):
    """
    Erstellt eine formatierte Excel-Datei aus den Strukturdaten
    
    Args:
        data (list): Liste von Dictionaries mit Strukturdaten
        output_path (str): Pfad zur Ausgabedatei
    """
    # DataFrame erstellen
    df = pd.DataFrame(data)
    
    # Einfache Ausgabe als Excel
    df.to_excel(output_path, index=False)
    
    # Excel-Datei mit openpyxl öffnen und formatieren
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 10  # Level
    ws.column_dimensions['B'].width = 15  # Symbol
    ws.column_dimensions['C'].width = 15  # Type
    ws.column_dimensions['D'].width = 30  # Title_de
    ws.column_dimensions['E'].width = 50  # Text_de
    
    # Überschriften formatieren
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
    
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Zellenränder für die gesamte Tabelle
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatierung für bestimmte Typen
    chapter_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    requirement_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Zeilen formatieren
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Typ aus Spalte C (index 2)
        cell_type = row[2].value
        
        # Füllung basierend auf Typ
        fill = chapter_fill if cell_type == "CHAPTER" else requirement_fill
        
        # Schriftgröße und Fett-Formatierung basierend auf Level
        level = row[0].value
        font_size = 12 if level == "1" else 11
        is_bold = level == "1" or level == "2"
        
        for cell in row:
            cell.border = thin_border
            cell.fill = fill
            cell.font = Font(size=font_size, bold=is_bold)
            
            # Text-Ausrichtung
            if cell.column == 1 or cell.column == 2 or cell.column == 3:  # Level, Symbol, Type
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Als Tabelle formatieren
    tab = Table(displayName="StrukturTabelle", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Filtern aktivieren
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    # Speichern
    wb.save(output_path)
    
    return output_path

if __name__ == "__main__":
    # Test-Daten
    test_data = [
        {"Level": "1", "Symbol": "A", "Type": "CHAPTER", "Title_de": "Einleitung", "Text_de": "Beschreibung der Einleitung"},
        {"Level": "2", "Symbol": "B1", "Type": "CHAPTER", "Title_de": "Definition", "Text_de": "Beschreibung der Definition"},
        {"Level": "3", "Symbol": "C2.1", "Type": "REQUIREMENT", "Title_de": "Anforderung", "Text_de": "Beschreibung der Anforderung"}
    ]
    
    create_excel_template(test_data, "template_example.xlsx")
    print("Excel-Template-Beispiel wurde erstellt: template_example.xlsx")
